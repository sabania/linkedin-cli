"""LinkedIn API wrapper using Selenium to bypass bot detection.

LinkedIn actively detects non-browser HTTP clients (PerimeterX/HUMAN Security)
and invalidates cookies. Using Selenium with a real Chrome instance ensures
all fingerprinting checks pass.
"""

import json
import time
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

CONFIG_DIR = Path.home() / ".linkedin-cli"
VOYAGER_API = "https://www.linkedin.com/voyager/api"


def _urn_to_timestamp(urn_or_id: str) -> str:
    """Extract timestamp from a LinkedIn Snowflake activity ID. Returns ISO 8601 UTC string."""
    from datetime import datetime, timezone
    try:
        numeric = int(str(urn_or_id).split(":")[-1])
        ms = (numeric >> 22) + 1197115136000
        return datetime.fromtimestamp(ms / 1000, tz=timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    except (ValueError, OSError):
        return ""


class LinkedinClient:
    """LinkedIn client using Selenium for all API calls."""

    def __init__(self, driver: webdriver.Chrome):
        self.driver = driver
        self._me_cache = None

    def _dismiss_modal(self):
        """Dismiss any cookie consent banner or overlay modal."""
        self.driver.execute_script(r'''
        // Cookie consent banner (artdeco-global-alert)
        var cookie = document.querySelector('.artdeco-global-alert--cookie_consent');
        if (cookie) {
            var btns = cookie.querySelectorAll('button');
            for (var i = 0; i < btns.length; i++) {
                var t = btns[i].innerText.toLowerCase();
                if (t.indexOf('akzeptier') >= 0 || t.indexOf('accept') >= 0 || t.indexOf('agree') >= 0) {
                    btns[i].click(); return;
                }
            }
        }
        // Standard modal dialog
        var modal = document.querySelector('.artdeco-modal:not(.vjs-modal-dialog)');
        if (modal) {
            var close = modal.querySelector('.artdeco-modal__dismiss');
            if (close) close.click();
        }
        ''')

    def _api_get(self, endpoint: str) -> dict:
        """Execute a Voyager API GET request via the browser."""
        url = f"{VOYAGER_API}{endpoint}"

        # Use JavaScript fetch() inside the real browser
        script = """
        const resp = await fetch(arguments[0], {
            headers: {
                'X-RestLi-Protocol-Version': '2.0.0',
                'csrf-token': document.cookie.match(/JSESSIONID="?([^";]+)/)?.[1] || '',
            },
            credentials: 'include',
        });
        return {status: resp.status, body: await resp.json()};
        """
        result = self.driver.execute_script(f"return (async () => {{ {script} }})()", url)

        if result["status"] != 200:
            return {}
        return result["body"]

    def get_user_profile(self, use_cache=True) -> dict:
        if self._me_cache and use_cache:
            return self._me_cache
        self._me_cache = self._api_get("/me")
        return self._me_cache

    def get_profile(self, public_id=None, urn_id=None) -> dict:
        identifier = public_id or urn_id
        if not identifier:
            return {}
        # Base profile data (name, headline, summary, etc.)
        data = self._api_get(
            f"/identity/dash/profiles?q=memberIdentity&memberIdentity={identifier}"
        )
        elements = data.get("elements", [])
        if not elements:
            return {}
        raw = elements[0]
        # Supplementary data (follower count, connection count)
        sup = self._api_get(
            f"/identity/dash/profiles?q=memberIdentity&memberIdentity={identifier}"
            f"&decorationId=com.linkedin.voyager.dash.deco.identity.profile.TopCardSupplementary-126"
        )
        sup_elements = sup.get("elements", [])
        if sup_elements:
            raw["followingState"] = sup_elements[0].get("followingState")
            raw["connections"] = sup_elements[0].get("connections")
        return _normalize_profile(raw)

    def get_profile_contact_info(self, public_id=None, urn_id=None) -> dict:
        identifier = public_id or urn_id
        if not identifier:
            return {}
        # Use dash profile which has the actual contact fields
        data = self._api_get(
            f"/identity/dash/profiles?q=memberIdentity&memberIdentity={identifier}"
        )
        elements = data.get("elements", [])
        if not elements:
            return {}
        p = elements[0]
        email_obj = p.get("emailAddress", {})
        website_obj = p.get("creatorWebsite", {})
        website_url = ""
        if website_obj:
            for attr in website_obj.get("attributesV2", []):
                href = attr.get("detailDataUnion", {}).get("hyperlink", "")
                if href:
                    website_url = href
                    break
        return {
            "email_address": email_obj.get("emailAddress", "") if isinstance(email_obj, dict) else "",
            "phone_numbers": [],
            "twitter": p.get("twitterHandles", []),
            "websites": [website_url] if website_url else [],
        }

    def get_profile_skills(self, public_id=None, urn_id=None) -> list:
        identifier = public_id or urn_id
        if not identifier:
            return []
        self.driver.get(f"https://www.linkedin.com/in/{identifier}/details/skills/")
        time.sleep(4)
        for i in range(5):
            self.driver.execute_script(f"window.scrollTo(0, {(i + 1) * 500})")
            time.sleep(0.3)
        time.sleep(1)
        script = r'''
        var main = document.querySelector('main');
        if (!main) return [];
        var skills = [];
        var seen = {};

        // Strategy 1: Edit links (own profile only)
        var editLinks = main.querySelectorAll('a[href*="/details/skills/edit/forms/"]');
        for (var i = 0; i < editLinks.length; i++) {
            var href = editLinks[i].getAttribute('href') || '';
            var formId = href.match(/edit\/forms\/(\w+)/);
            if (!formId || seen[formId[1]]) continue;
            seen[formId[1]] = true;
            var parent = editLinks[i].parentElement;
            if (!parent) continue;
            var prev = parent.previousElementSibling;
            if (prev) {
                var name = prev.innerText.trim().split('\n')[0].trim();
                if (name && name.length > 0 && name.length < 80) {
                    skills.push({name: name});
                }
            }
        }

        // Strategy 2: Third-party profile — parse innerText (language-independent)
        if (skills.length === 0) {
            var text = main.innerText || '';
            var lines = text.split('\n').map(function(l) { return l.trim(); }).filter(function(l) { return l.length > 0; });

            // Find content start: the first skill is the first line followed by an
            // endorsement count (digit + space). This is language-independent.
            var contentStart = 0;
            for (var s = 0; s < Math.min(lines.length, 15); s++) {
                if (s + 1 < lines.length && /^\d+\s/.test(lines[s + 1])) {
                    contentStart = s;
                    break;
                }
            }
            // Fallback: skip title + up to 4 tabs (5 lines max)
            if (contentStart === 0) {
                contentStart = 1;
                while (contentStart < Math.min(lines.length, 5) && lines[contentStart].length < 35) {
                    contentStart++;
                }
            }

            // Find footer: degree indicator "·N." marks suggested profiles section
            var footerStart = lines.length;
            for (var f = contentStart; f < lines.length; f++) {
                if (/^[\u00B7·]\s*\d/.test(lines[f])) {
                    footerStart = f;
                    break;
                }
            }

            // Collect skills up to footer
            for (var i = contentStart; i < footerStart; i++) {
                var line = lines[i];
                // Skip endorsement count lines (start with digit + space)
                if (/^\d+\s/.test(line)) continue;
                // Skip very short or very long
                if (line.length <= 1 || line.length > 80) continue;
                if (!seen[line]) {
                    seen[line] = true;
                    skills.push({name: line});
                }
            }

            // Clean trailing false positives: remove entries from the last 3 lines
            // before footer (suggested profile name, "people also viewed" header, buttons)
            if (footerStart < lines.length) {
                while (skills.length > 0) {
                    var lastName = skills[skills.length - 1].name;
                    var isFalse = false;
                    for (var c = Math.max(contentStart, footerStart - 3); c < footerStart; c++) {
                        if (lines[c] === lastName) { isFalse = true; break; }
                    }
                    if (isFalse) skills.pop(); else break;
                }
            }
        }

        return skills;
        '''
        return self.driver.execute_script(script) or []

    def get_profile_posts(self, public_id=None, urn_id=None, limit=25, post_count=None) -> list:
        """Get posts from a profile via GraphQL API with pagination."""
        import re as _re
        limit = post_count or limit
        identifier = public_id or urn_id
        if not identifier:
            return []

        # Load activity page to capture the GraphQL queryId
        self.driver.get(f"https://www.linkedin.com/in/{identifier}/recent-activity/all/")
        time.sleep(4)

        urls = self.driver.execute_script(r'''
        var entries = performance.getEntriesByType("resource");
        for (var i = 0; i < entries.length; i++) {
            if (entries[i].name.indexOf("ProfileUpdates") >= 0) return entries[i].name;
        }
        return null;
        ''')
        if not urls:
            return []

        qid_match = _re.search(r'queryId=([^&]+)', urls)
        profile_urn_match = _re.search(r'profileUrn:([^)]+)', urls)
        if not qid_match or not profile_urn_match:
            return []

        query_id = qid_match.group(1)
        profile_urn = profile_urn_match.group(1)

        all_posts = []
        page_size = 20
        start = 0

        while start < limit:
            count = min(page_size, limit - start)
            data = self._api_get(
                f"/graphql?variables=(count:{count},start:{start},profileUrn:{profile_urn})&queryId={query_id}"
            )
            feed = (data or {}).get("data", {})
            # Find the key containing elements
            elements = []
            for k, v in feed.items():
                if isinstance(v, dict) and "elements" in v:
                    elements = v["elements"]
                    break
            if not elements:
                break

            for el in elements:
                entity_urn = el.get("entityUrn", "")
                activity_match = _re.search(r'urn:li:activity:\d+', entity_urn)
                urn = activity_match.group(0) if activity_match else entity_urn

                commentary = el.get("commentary") or {}
                text_obj = commentary.get("text") or {}
                text = text_obj.get("text", "") if isinstance(text_obj, dict) else ""

                social = el.get("socialDetail") or {}
                counts = social.get("totalSocialActivityCounts") or {}

                all_posts.append({
                    "urn": urn,
                    "text": text[:500],
                    "reactions": str(counts.get("numLikes", 0)),
                    "comments": str(counts.get("numComments", 0)),
                    "shares": str(counts.get("numShares", 0)),
                    "posted_at": _urn_to_timestamp(urn),
                    "post_url": f"https://www.linkedin.com/feed/update/{urn}/" if urn else "",
                })

            start += len(elements)
            if len(elements) < count:
                break

        return all_posts[:limit]

    def get_feed_posts(self, limit=25, exclude_promoted_posts=True) -> list:
        """Get feed posts via GraphQL API with pagination."""
        # Load feed page to capture the GraphQL queryId
        self.driver.get("https://www.linkedin.com/feed/")
        time.sleep(4)

        import re
        feed_url = self.driver.execute_script(r'''
        var entries = performance.getEntriesByType("resource");
        for (var i = 0; i < entries.length; i++) {
            if (entries[i].name.indexOf("MainFeed") >= 0) return entries[i].name;
        }
        return null;
        ''')
        if not feed_url:
            return []

        qid_match = re.search(r'queryId=([^&]+)', feed_url)
        if not qid_match:
            return []
        query_id = qid_match.group(1)

        all_posts = []
        page_size = 25
        start = 0

        while start < limit:
            count = min(page_size, limit - start)
            data = self._api_get(
                f"/graphql?variables=(start:{start},count:{count},sortOrder:RELEVANCE)&queryId={query_id}"
            )
            feed = (data or {}).get("data", {}).get("feedDashMainFeedByMainFeed", {})
            elements = feed.get("elements", [])
            if not elements:
                break

            for el in elements:
                # Extract activity URN
                entity_urn = el.get("entityUrn", "")
                activity_match = re.search(r'urn:li:activity:\d+', entity_urn)
                urn = activity_match.group(0) if activity_match else entity_urn

                # Text
                commentary = el.get("commentary") or {}
                text_obj = commentary.get("text") or {}
                text = text_obj.get("text", "") if isinstance(text_obj, dict) else ""

                # Author
                actor = el.get("actor") or {}
                name_obj = actor.get("name") or {}
                author = name_obj.get("text", "") if isinstance(name_obj, dict) else str(name_obj)
                # Extract author profile ID from navigationContext.actionTarget
                author_profile_id = ""
                nav_ctx = actor.get("navigationContext") or {}
                action_target = nav_ctx.get("actionTarget", "") if isinstance(nav_ctx, dict) else ""
                if "/in/" in action_target:
                    author_profile_id = action_target.split("/in/")[-1].split("?")[0].rstrip("/")

                # Social counts
                social = el.get("socialDetail") or {}
                counts = social.get("totalSocialActivityCounts") or {}
                reactions = str(counts.get("numLikes", 0))
                comments = str(counts.get("numComments", 0))

                # Repost detection
                is_repost = bool(el.get("resharedUpdate"))

                all_posts.append({
                    "urn": urn,
                    "text": text[:500],
                    "author": author,
                    "author_profile_id": author_profile_id,
                    "reactions": reactions,
                    "comments": comments,
                    "shares": str(counts.get("numShares", 0)),
                    "posted_at": _urn_to_timestamp(urn),
                    "post_url": f"https://www.linkedin.com/feed/update/{urn}/" if urn else "",
                    "is_repost": is_repost,
                })

            start += len(elements)
            if len(elements) < count:
                break

        return all_posts[:limit]

    def _get_post_analytics_list(self, activity_id: str, result_type: str, limit: int) -> list:
        """Get reactions/comments/reposts via the analytics GraphQL API with pagination."""
        # Load analytics page to capture queryId
        self.driver.get(
            f"https://www.linkedin.com/analytics/post/urn:li:activity:{activity_id}/?resultType={result_type}"
        )
        time.sleep(4)

        import re as _re
        qid_url = self.driver.execute_script(r'''
        var entries = performance.getEntriesByType("resource");
        for (var i = 0; i < entries.length; i++) {
            if (entries[i].name.indexOf("AnalyticsObject") >= 0) return entries[i].name;
        }
        return null;
        ''')

        if not qid_url:
            return []
        qid = _re.search(r'queryId=([^&]+)', qid_url).group(1)
        act_urn = f"urn%3Ali%3Aactivity%3A{activity_id}"

        all_items = []
        page_size = 10
        start = 0

        while start < limit:
            data = self._api_get(
                f"/graphql?variables=(start:{start},"
                f"query:(selectedFilters:List((key:resultType,value:List({result_type})))),"
                f"analyticsEntityUrn:(activityUrn:{act_urn}),surfaceType:POST)"
                f"&queryId={qid}"
            )
            feed = (data or {}).get("data", {})
            elements = []
            for k, v in feed.items():
                if isinstance(v, dict) and "elements" in v:
                    elements = v["elements"]
                    break
            if not elements:
                break

            for el in elements:
                entity = (el.get("content") or {}).get("analyticsEntityLockup", {}).get("entityLockup", {})
                title = entity.get("title") or {}
                subtitle = entity.get("subtitle") or {}
                nav_url = entity.get("navigationUrl", "")
                # Extract public ID from URL
                pid = nav_url.split("/in/")[-1].rstrip("/") if "/in/" in nav_url else ""

                all_items.append({
                    "name": title.get("text", ""),
                    "headline": (subtitle.get("text", "") or "")[:100],
                    "profileId": pid,
                    "profileUrl": nav_url,
                })

            start += len(elements)
            if len(elements) < page_size:
                break

        return all_items[:limit]

    def get_post_comments(self, post_urn: str, limit=50, comment_count=None) -> list:
        """Get comments on a post by loading the post page (scraping — analytics API has no comment text)."""
        limit = comment_count or limit
        activity_id = post_urn.split(":")[-1]
        self.driver.get(f"https://www.linkedin.com/feed/update/urn:li:activity:{activity_id}/")
        time.sleep(4)

        # Click "load more comments" buttons
        click_count = min(max(3, limit // 10), 15)
        for _ in range(click_count):
            try:
                found = self.driver.execute_script(
                    "var btn = document.querySelector('[class*=\"show-prev\"], button[class*=\"comments-comments-list__load-more\"]');"
                    "if (btn) { btn.click(); return true; } return false;"
                )
                if not found:
                    break
                time.sleep(1)
            except Exception:
                break

        script = (
            "var comments = [];"
            "var els = document.querySelectorAll('[class*=\"comments-comment-item\"], [class*=\"comment-item\"]');"
            "els.forEach(function(el) {"
            "  var textEl = el.querySelector('[class*=\"comment-item__main-content\"] span[dir], [class*=\"comment__text\"] span[dir]');"
            "  if (!textEl) { var spans = el.querySelectorAll('span[dir]'); for (var i=0;i<spans.length;i++) { if (spans[i].innerText.length > 20) { textEl = spans[i]; break; } } }"
            "  var link = el.querySelector('a[href*=\"/in/\"]');"
            "  var name = '';"
            "  if (link) {"
            "    var nameEl = link.querySelector('span[dir] span, span.hoverable-link-text, span');"
            "    if (nameEl) name = nameEl.innerText.trim().split('\\n')[0];"
            "    if (!name) name = link.innerText.trim().split('\\n')[0];"
            "  }"
            "  if (!name) {"
            "    var actorEl = el.querySelector('[class*=\"comment-item__actor\"], [class*=\"actor\"]');"
            "    if (actorEl) name = actorEl.innerText.trim().split('\\n')[0];"
            "  }"
            "  comments.push({"
            "    author: name,"
            "    text: textEl ? textEl.innerText.trim().substring(0, 500) : '',"
            "    profileUrl: link ? link.getAttribute('href') : '',"
            "    profileId: link ? link.getAttribute('href').split('/in/')[1].replace('/','') : ''"
            "  });"
            "});"
            "return comments;"
        )
        comments = self.driver.execute_script(script) or []
        seen = set()
        unique = []
        for c in comments:
            key = c.get("text", "")[:50]
            if key and key not in seen:
                seen.add(key)
                unique.append(c)
        return unique[:limit]

    def get_post_reactions(self, post_urn: str, limit=50, max_results=None) -> list:
        """Get reactors on a post via analytics API."""
        count = max_results or limit
        activity_id = post_urn.split(":")[-1]
        return self._get_post_analytics_list(activity_id, "REACTIONS", count)

    def _extract_search_results(self, data: dict) -> list:
        """Extract entity results from search cluster response."""
        results = []
        for cluster in data.get("elements", []):
            for item in cluster.get("items", []):
                entity = (
                    item.get("itemUnion", {}).get("entityResult")
                    or item.get("item", {}).get("entityResult")
                    or {}
                )
                if not entity:
                    continue
                title = entity.get("title", {}).get("text", "")
                subtitle = entity.get("primarySubtitle", {}).get("text", "")
                secondary = entity.get("secondarySubtitle", {}).get("text", "") if entity.get("secondarySubtitle") else ""
                nav_url = entity.get("navigationUrl", "")
                urn = entity.get("entityUrn", "")
                public_id = nav_url.split("/in/")[-1].split("?")[0].rstrip("/") if "/in/" in nav_url else ""
                badge_obj = entity.get("badgeText") or {}
                badge = badge_obj.get("text", "") if isinstance(badge_obj, dict) else ""
                results.append({
                    "name": title,
                    "headline": subtitle,
                    "location": secondary,
                    "public_id": public_id,
                    "urn_id": urn,
                    "url": nav_url.split("?")[0] if nav_url else "",
                    "connection_degree": badge,
                })
        return results

    def _resolve_public_ids(self, urns: list) -> dict:
        """Resolve a list of URN-based member IDs to public profile IDs in parallel."""
        script = """
        const urns = arguments[0];
        const csrf = document.cookie.match(/JSESSIONID="?([^";]+)/)?.[1] || '';
        const results = await Promise.all(urns.map(async (urn) => {
            try {
                const resp = await fetch(
                    'https://www.linkedin.com/voyager/api/identity/dash/profiles?q=memberIdentity&memberIdentity=' + urn,
                    { headers: {'X-RestLi-Protocol-Version':'2.0.0','csrf-token':csrf}, credentials:'include' }
                );
                if (resp.status !== 200) return {urn, publicId: urn};
                const data = await resp.json();
                const el = data.elements && data.elements[0];
                return {urn, publicId: (el && el.publicIdentifier) || urn};
            } catch(e) { return {urn, publicId: urn}; }
        }));
        return results;
        """
        resolved = self.driver.execute_script(
            f"return (async () => {{ {script} }})()", urns
        ) or []
        return {r["urn"]: r["publicId"] for r in resolved}

    def get_current_profile_views(self) -> list:
        """Get list of people who viewed your profile."""
        data = self._api_get("/identity/wvmpCards")
        return data.get("elements", [])

    def get_profile_connections(self, public_id=None, urn_id=None, max_results=50) -> list:
        """Get connections of a profile. Accepts public_id or urn_id."""
        # connectionOf needs the member identity hash
        if urn_id:
            member_id = urn_id.split(":")[-1] if ":" in urn_id else urn_id
        elif public_id:
            # Resolve public ID to URN via profile lookup
            profile = self.get_profile(public_id=public_id)
            urn = profile.get("entityUrn", "")
            member_id = urn.split(":")[-1] if urn else ""
            if not member_id:
                return []
        else:
            return []
        results = []
        start = 0
        page_size = 49  # LinkedIn max per page
        while start < max_results:
            count = min(page_size, max_results - start)
            data = self._api_get(
                f"/search/dash/clusters?decorationId=com.linkedin.voyager.dash.deco.search.SearchClusterCollection-175"
                f"&origin=MEMBER_PROFILE_CANNED_SEARCH&q=all&start={start}&count={count}"
                f"&query=(flagshipSearchIntent:SEARCH_SRP,queryParameters:"
                f"(resultType:List(PEOPLE),network:List(F),connectionOf:List({member_id})))"
            )
            page_results = []
            for cluster in data.get("elements", []):
                for item in cluster.get("items", []):
                    entity = (
                        item.get("itemUnion", {}).get("entityResult")
                        or item.get("item", {}).get("entityResult")
                        or {}
                    )
                    if not entity:
                        continue
                    title = entity.get("title", {}).get("text", "")
                    subtitle = entity.get("primarySubtitle", {}).get("text", "")
                    nav_url = entity.get("navigationUrl", "")
                    pid = nav_url.split("/in/")[-1].split("?")[0].rstrip("/") if "/in/" in nav_url else ""
                    page_results.append({
                        "firstName": title.split(" ")[0] if title else "",
                        "lastName": " ".join(title.split(" ")[1:]) if title else "",
                        "headline": subtitle,
                        "public_id": pid,
                    })
            results.extend(page_results)
            if len(page_results) < count:
                break  # no more results
            start += len(page_results)
        return results

    def get_profile_experiences(self, public_id: str = None, urn_id: str = None) -> list:
        """Get work experiences of a profile by scraping the experience section."""
        profile_id = public_id or urn_id
        self.driver.get(f"https://www.linkedin.com/in/{profile_id}/details/experience/")
        time.sleep(4)
        for i in range(8):
            self.driver.execute_script(f"window.scrollTo(0, {(i + 1) * 600})")
            time.sleep(0.3)
        time.sleep(1)
        script = r'''
        var main = document.querySelector('main');
        if (!main) return [];
        var exps = [];
        var midDot = '\u00B7';

        function parseExpLines(lines) {
            if (lines.length === 0) return null;
            var title = lines[0] || '';
            var companyName = '';
            var period = '';
            var location = '';
            for (var j = 1; j < lines.length; j++) {
                var line = lines[j];
                if (!companyName && line.indexOf(midDot) >= 0 && !/\d{4}/.test(line.split(midDot)[0])) {
                    companyName = line.split(midDot)[0].trim();
                } else if (!period && /\d{4}/.test(line) && (line.indexOf('\u2013') >= 0 || line.indexOf('-') >= 0 || line.indexOf(midDot) >= 0)) {
                    period = line;
                } else if (period && !location && line.length < 80) {
                    location = line;
                }
            }
            return {title: title, companyName: companyName, timePeriod: period, location: location};
        }

        function findCompanyFromLogo(startEl) {
            var el = startEl;
            for (var d = 0; d < 12; d++) {
                el = el.parentElement;
                if (!el || el === main) break;
                if (el.tagName.toLowerCase() === 'ul') {
                    var groupDiv = el.parentElement;
                    if (groupDiv) {
                        var img = groupDiv.querySelector('img[alt]');
                        if (img) {
                            var alt = img.getAttribute('alt') || '';
                            return alt.replace(/^(Logo|Logotipo|Logotype)\s+(von|of|de|di|du|van)\s+/i, '')
                                      .replace(/\s+(logo|Logo)$/i, '').trim();
                        }
                    }
                    break;
                }
            }
            return '';
        }

        // Strategy 1: Edit links (own profile only)
        var editLinks = main.querySelectorAll('a[href*="/details/experience/edit/forms/"]');
        var seen = {};

        for (var i = 0; i < editLinks.length; i++) {
            var link = editLinks[i];
            var href = link.getAttribute('href') || '';
            var formId = href.match(/edit\/forms\/(\d+)/);
            if (!formId || seen[formId[1]]) continue;
            seen[formId[1]] = true;

            var linkText = link.innerText.trim();
            var lines = linkText.split('\n').filter(function(l) { return l.trim().length > 0; }).map(function(l) { return l.trim(); });
            var exp = parseExpLines(lines);
            if (!exp) continue;
            if (!exp.companyName) exp.companyName = findCompanyFromLogo(link);
            exps.push(exp);
        }

        // Strategy 2: Third-party profile — parse innerText (language-independent)
        if (exps.length === 0) {
            var text = main.innerText || '';
            var allLines = text.split('\n').map(function(l) { return l.trim(); }).filter(function(l) { return l.length > 0; });
            var yearRe = /\b(19|20)\d{2}\b/;

            for (var idx = 2; idx < allLines.length; idx++) {
                var ln = allLines[idx];
                // Stop at connection degree indicators (universal footer marker: "·N.")
                if (/^[\u00B7·]\s*\d/.test(ln)) break;

                // Detect period line (language-independent):
                // Must have: 4-digit year + midDot (duration separator) + en-dash (date range)
                // LinkedIn universally uses en-dash \u2013 and midDot \u00B7 regardless of UI language
                var isPeriod = yearRe.test(ln) && ln.indexOf(midDot) >= 0 &&
                    (ln.indexOf('\u2013') >= 0 || ln.indexOf('\u2014') >= 0);
                if (!isPeriod) continue;

                // Company is the line before period (usually has midDot for employment type)
                var compLine = allLines[idx - 1] || '';
                var comp = '';
                if (compLine.indexOf(midDot) >= 0 && !yearRe.test(compLine.split(midDot)[0])) {
                    comp = compLine.split(midDot)[0].trim();
                } else {
                    comp = compLine;
                }

                // Title is 2 lines before period (1 before company)
                var ttl = (idx >= 2) ? allLines[idx - 2] : '';

                // Skip if title looks like a skills summary (universal: ends with "+N word")
                if (!ttl || ttl.length === 0 || ttl.length > 100) continue;
                if (/\+\d+\s+\S+$/.test(ttl)) continue;

                // Location is the line after period (short, no year)
                var loc = '';
                if (idx + 1 < allLines.length) {
                    var nxt = allLines[idx + 1];
                    if (nxt.length < 80 && !yearRe.test(nxt)) {
                        loc = nxt;
                    }
                }

                var key = ttl + '|' + comp;
                if (!seen[key]) {
                    seen[key] = true;
                    exps.push({title: ttl, companyName: comp, timePeriod: ln, location: loc});
                }
            }
        }

        return exps;
        '''
        return self.driver.execute_script(script) or []

    def get_profile_network_info(self, public_profile_id: str) -> dict:
        """Get network info (follower count, connection count)."""
        data = self._api_get(
            f"/identity/profiles/{public_profile_id}/networkinfo"
        )
        return data

    def _get_messaging_query_ids(self) -> dict:
        """Capture messaging GraphQL queryIds and mailboxUrn from the messaging page."""
        import re as _re
        urls = self.driver.execute_script(r'''
        var entries = performance.getEntriesByType("resource");
        var result = {};
        for (var i = 0; i < entries.length; i++) {
            var u = entries[i].name;
            var qid = u.match(/queryId=([^&]+)/);
            if (!qid) continue;
            if (u.indexOf("messengerConversations") >= 0) {
                result.conversations = qid[1];
                var mb = u.match(/mailboxUrn:([^)&]+)/);
                if (mb) result.mailboxUrn = decodeURIComponent(mb[1]);
            }
            if (u.indexOf("messengerMessages") >= 0) result.messages = qid[1];
        }
        return result;
        ''') or {}
        return urls

    def get_conversations(self, limit=25) -> list:
        """Get conversations via GraphQL messaging API."""
        self.driver.get("https://www.linkedin.com/messaging/")
        time.sleep(4)

        qids = self._get_messaging_query_ids()
        conv_qid = qids.get("conversations")
        my_urn = qids.get("mailboxUrn", "")
        if not conv_qid or not my_urn:
            return []
        my_urn_encoded = my_urn.replace(":", "%3A")

        data = self._api_get(
            f"/voyagerMessagingGraphQL/graphql?queryId={conv_qid}"
            f"&variables=(mailboxUrn:{my_urn_encoded})"
        )
        feed = (data or {}).get("data", {})
        # Find the key with elements
        elements = []
        for k, v in feed.items():
            if isinstance(v, dict) and "elements" in v:
                elements = v["elements"]
                break

        convos = []
        for el in elements[:limit]:
            # Get participant names (skip self)
            parts = el.get("conversationParticipants", [])
            names = []
            for p in parts:
                if my_urn in p.get("entityUrn", ""):
                    continue
                pt = (p.get("participantType") or {})
                member = (pt.get("member") or {})
                fn = member.get("firstName") or {}
                ln = member.get("lastName") or {}
                fn_text = fn.get("text", "") if isinstance(fn, dict) else str(fn)
                ln_text = ln.get("text", "") if isinstance(ln, dict) else str(ln)
                full = f"{fn_text} {ln_text}".strip()
                if full:
                    names.append(full)

            # Get last message preview
            msgs = el.get("messages") or {}
            msg_els = msgs.get("elements", []) if isinstance(msgs, dict) else []
            last_text = ""
            if msg_els:
                body = msg_els[0].get("body") or {}
                last_text = body.get("text", "") if isinstance(body, dict) else str(body)

            # Conversation URN for read_conversation
            entity_urn = el.get("entityUrn", "")

            from datetime import datetime
            ts = el.get("lastActivityAt", 0)
            date_str = datetime.fromtimestamp(ts / 1000).strftime("%Y-%m-%d %H:%M") if ts else ""

            convos.append({
                "participants": ", ".join(names) or "Unknown",
                "lastMessage": last_text[:200],
                "date": date_str,
                "conversationUrn": entity_urn,
            })

        return convos[:limit]

    def get_conversation(self, conversation_urn_id: str = None, name: str = None) -> list:
        """Get messages from a conversation via GraphQL API."""
        if name:
            # Load conversations and find matching one
            convos = self.get_conversations(limit=50)
            target = name.lower()
            conv_urn = None
            for c in convos:
                if target in c.get("participants", "").lower():
                    conv_urn = c.get("conversationUrn", "")
                    break
            if not conv_urn:
                return []
            return self._get_messages_by_urn(conv_urn)
        elif conversation_urn_id:
            # Build full conversation URN if just the thread ID was given
            if not conversation_urn_id.startswith("urn:"):
                me = self.get_user_profile()
                my_urn = me.get("entityUrn", "")
                conversation_urn_id = f"urn:li:msg_conversation:({my_urn},{conversation_urn_id})"
            return self._get_messages_by_urn(conversation_urn_id)
        return []

    def _get_messages_by_urn(self, conversation_urn: str) -> list:
        """Fetch messages for a conversation URN via GraphQL."""
        # Ensure messaging page is loaded (for queryId capture)
        if "messaging" not in (self.driver.current_url or ""):
            self.driver.get("https://www.linkedin.com/messaging/")
            time.sleep(4)

        qids = self._get_messaging_query_ids()
        msg_qid = qids.get("messages")
        if not msg_qid:
            return []

        conv_urn_encoded = conversation_urn.replace(":", "%3A").replace("(", "%28").replace(")", "%29").replace(",", "%2C")

        data = self._api_get(
            f"/voyagerMessagingGraphQL/graphql?queryId={msg_qid}"
            f"&variables=(conversationUrn:{conv_urn_encoded})"
        )
        feed = (data or {}).get("data", {})
        elements = []
        for k, v in feed.items():
            if isinstance(v, dict) and "elements" in v:
                elements = v["elements"]
                break

        msgs = []
        for msg in elements:
            body = msg.get("body") or {}
            text = body.get("text", "") if isinstance(body, dict) else str(body)

            sender_obj = msg.get("sender") or msg.get("actor") or {}
            pt = (sender_obj.get("participantType") or {})
            member = (pt.get("member") or {})
            fn = member.get("firstName") or {}
            ln = member.get("lastName") or {}
            fn_text = fn.get("text", "") if isinstance(fn, dict) else str(fn)
            ln_text = ln.get("text", "") if isinstance(ln, dict) else str(ln)
            sender_name = f"{fn_text} {ln_text}".strip()

            from datetime import datetime
            ts = msg.get("deliveredAt", 0)
            time_str = datetime.fromtimestamp(ts / 1000).strftime("%Y-%m-%d %H:%M") if ts else ""

            if text:
                msgs.append({"sender": sender_name, "body": text, "time": time_str})

        return msgs

    def send_message(self, message_body: str, conversation_urn_id=None, recipients=None):
        # Messages require POST - use fetch in browser
        if conversation_urn_id:
            url = f"{VOYAGER_API}/messaging/conversations/{conversation_urn_id}/events"
        else:
            url = f"{VOYAGER_API}/messaging/conversations"

        payload = {"eventCreate": {"value": {"com.linkedin.voyager.messaging.create.MessageCreate": {"body": message_body}}}}
        if recipients:
            payload["recipients"] = recipients

        script = """
        const resp = await fetch(arguments[0], {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json',
                'X-RestLi-Protocol-Version': '2.0.0',
                'csrf-token': document.cookie.match(/JSESSIONID="?([^";]+)/)?.[1] || '',
            },
            credentials: 'include',
            body: JSON.stringify(arguments[1]),
        });
        return {status: resp.status};
        """
        return self.driver.execute_script(
            f"return (async () => {{ {script} }})()", url, payload
        )

    def mark_conversation_as_seen(self, conversation_urn_id: str):
        """Mark a conversation as seen."""
        url = f"{VOYAGER_API}/messaging/conversations/{conversation_urn_id}"
        script = """
        const resp = await fetch(arguments[0], {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json',
                'X-RestLi-Protocol-Version': '2.0.0',
                'csrf-token': document.cookie.match(/JSESSIONID="?([^";]+)/)?.[1] || '',
                'X-HTTP-Method-Override': 'PATCH',
            },
            credentials: 'include',
            body: JSON.stringify({patch: {$set: {read: true}}}),
        });
        return {status: resp.status};
        """
        return self.driver.execute_script(
            f"return (async () => {{ {script} }})()", url
        )

    def get_invitations(self, limit=25) -> list:
        from datetime import datetime, timezone
        raw_results = []
        page_size = min(limit, 49)
        start = 0
        while start < limit:
            count = min(page_size, limit - start)
            data = self._api_get(f"/relationships/invitationViews?start={start}&count={count}")
            page = data.get("elements", [])
            raw_results.extend(page)
            if len(page) < count:
                break
            start += len(page)

        normalized = []
        for inv in raw_results[:limit]:
            from_member = inv.get("fromMember", {})
            fn = from_member.get("firstName", "")
            ln = from_member.get("lastName", "")
            headline = from_member.get("occupation", "") or from_member.get("headline", "")
            pub_id = from_member.get("publicIdentifier", "")
            sent_ts = inv.get("sentTime", 0)
            sent_time = datetime.fromtimestamp(sent_ts / 1000, tz=timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ") if sent_ts else ""
            normalized.append({
                "name": f"{fn} {ln}".strip(),
                "headline": headline,
                "profile_id": pub_id,
                "message": inv.get("message", ""),
                "entity_urn": inv.get("entityUrn", ""),
                "shared_secret": inv.get("sharedSecret", ""),
                "sent_time": sent_time,
            })
        return normalized

    def add_connection(self, profile_public_id: str, message="", profile_urn=None):
        import base64
        import os
        # Resolve public ID to profile URN
        profile = self.get_profile(public_id=profile_public_id)
        urn = profile.get("entityUrn", "")
        if not urn:
            return {"status": 404}
        url = f"{VOYAGER_API}/relationships/dash/memberRelationships?action=verifyQuotaAndCreate"
        tracking_id = base64.b64encode(os.urandom(16)).decode()
        payload = {
            "inviteeProfileUrn": urn,
            "trackingId": tracking_id,
        }

        script = """
        const resp = await fetch(arguments[0], {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json',
                'X-RestLi-Protocol-Version': '2.0.0',
                'csrf-token': document.cookie.match(/JSESSIONID="?([^";]+)/)?.[1] || '',
            },
            credentials: 'include',
            body: JSON.stringify(arguments[1]),
        });
        return {status: resp.status};
        """
        return self.driver.execute_script(
            f"return (async () => {{ {script} }})()", url, payload
        )

    def remove_connection(self, public_profile_id: str):
        """Remove a connection."""
        url = f"{VOYAGER_API}/identity/profiles/{public_profile_id}/profileActions?action=disconnect"
        script = """
        const resp = await fetch(arguments[0], {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json',
                'X-RestLi-Protocol-Version': '2.0.0',
                'csrf-token': document.cookie.match(/JSESSIONID="?([^";]+)/)?.[1] || '',
            },
            credentials: 'include',
        });
        return {status: resp.status};
        """
        return self.driver.execute_script(
            f"return (async () => {{ {script} }})()", url
        )

    def reply_invitation(self, invitation_entity_urn: str, invitation_shared_secret: str, action: str = "accept"):
        """Accept or reject an invitation."""
        url = f"{VOYAGER_API}/relationships/invitations/{invitation_entity_urn}?action={action}"
        payload = {"invitationSharedSecret": invitation_shared_secret}
        script = """
        const resp = await fetch(arguments[0], {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json',
                'X-RestLi-Protocol-Version': '2.0.0',
                'csrf-token': document.cookie.match(/JSESSIONID="?([^";]+)/)?.[1] || '',
            },
            credentials: 'include',
            body: JSON.stringify(arguments[1]),
        });
        return {status: resp.status};
        """
        return self.driver.execute_script(
            f"return (async () => {{ {script} }})()", url, payload
        )

    def search_people(self, keywords=None, limit=25, **kwargs) -> list:
        kw = keywords or ""
        # Merge keyword-based filters into the keywords string
        for k in ("keyword_company", "keyword_title", "keyword_school"):
            if kwargs.get(k):
                kw = f"{kw} {kwargs[k]}".strip()

        # Build dynamic queryParameters
        qp_parts = ["resultType:List(PEOPLE)"]
        if kwargs.get("network_depths"):
            qp_parts.append(f"network:List({','.join(kwargs['network_depths'])})")
        if kwargs.get("regions"):
            qp_parts.append(f"geoUrn:List({','.join(kwargs['regions'])})")
        qp_str = ",".join(qp_parts)

        results = []
        page_size = min(limit, 49)
        start = 0
        while start < limit:
            count = min(page_size, limit - start)
            data = self._api_get(
                f"/search/dash/clusters?decorationId=com.linkedin.voyager.dash.deco.search.SearchClusterCollection-175"
                f"&origin=GLOBAL_SEARCH_HEADER&q=all&start={start}&count={count}"
                f"&query=(keywords:{kw},flagshipSearchIntent:SEARCH_SRP"
                f",queryParameters:({qp_str}))"
            )
            page = self._extract_search_results(data)
            results.extend(page)
            if len(page) < count:
                break
            start += len(page)
        return results[:limit]

    def get_company(self, public_id) -> dict:
        data = self._api_get(f"/organization/companies?q=universalName&universalName={public_id}")
        elements = data.get("elements", [])
        return elements[0] if elements else {}

    def get_company_updates(self, public_id=None, urn_id=None, limit=25, max_results=None) -> list:
        limit = max_results or limit
        identifier = public_id or urn_id
        # Scrape company posts page since the API endpoint is deprecated
        self.driver.get(f"https://www.linkedin.com/company/{identifier}/posts/")
        time.sleep(4)
        scroll_count = min(max(3, limit // 3), 15)
        for i in range(scroll_count):
            self.driver.execute_script(f"window.scrollBy(0, 1000)")
            time.sleep(0.7)
        script = (
            "var posts = [];"
            "var els = document.querySelectorAll('[data-urn]');"
            "els.forEach(function(el) {"
            "  var urn = el.getAttribute('data-urn');"
            "  if (!urn || urn.indexOf('activity') === -1) return;"
            "  var text = '';"
            "  var spans = el.querySelectorAll('span[dir]');"
            "  for (var i = 0; i < spans.length; i++) {"
            "    var t = spans[i].innerText || '';"
            "    if (t.length > text.length && t.length > 20) text = t;"
            "  }"
            "  var rxBtn = el.querySelector('button[data-reaction-details]');"
            "  var rxText = rxBtn ? rxBtn.innerText.trim() : '';"
            "  var rxMatch = rxText.match(/(\\d[\\d.,]*)/);"
            "  var rxCount = rxMatch ? rxMatch[1] : (rxBtn ? '1' : '0');"
            "  var cmLi = el.querySelector('li.social-details-social-counts__comments');"
            "  var cmBtn = cmLi ? cmLi.querySelector('button') : null;"
            "  var cmText = cmBtn ? cmBtn.innerText.trim() : '';"
            "  var cmMatch = cmText.match(/(\\d[\\d.,]*)/);"
            "  var cmCount = cmMatch ? cmMatch[1] : '0';"
            "  posts.push({urn: urn, text: text.substring(0, 500), reactions: rxCount, comments: cmCount});"
            "});"
            "return posts;"
        )
        posts = self.driver.execute_script(script) or []
        # Enrich with timestamps and URLs
        for p in posts:
            urn = p.get("urn", "")
            p["posted_at"] = _urn_to_timestamp(urn)
            p["post_url"] = f"https://www.linkedin.com/feed/update/{urn}/" if urn else ""
        return posts[:limit]

    def follow_company(self, public_id: str = None, following_state_urn: str = None, following: bool = True):
        """Follow or unfollow a company by clicking the button on the company page."""
        company_id = public_id or following_state_urn
        self.driver.get(f"https://www.linkedin.com/company/{company_id}/")
        time.sleep(4)
        # Find follow button by its CSS class (language-independent)
        script = r'''
        var btn = document.querySelector('.org-company-follow-button, button[class*="follow-button"]');
        if (!btn) return {clicked: false, error: 'no follow button found'};
        var isFollowing = btn.classList.contains('is-following') ||
                          btn.getAttribute('aria-pressed') === 'true';
        var wantFollow = arguments[0];
        if (wantFollow && !isFollowing) {
            btn.click();
            return {clicked: true, action: 'followed'};
        } else if (!wantFollow && isFollowing) {
            btn.click();
            return {clicked: true, action: 'unfollowed', needsConfirm: true};
        }
        return {clicked: false, action: wantFollow ? 'already following' : 'not following'};
        '''
        result = self.driver.execute_script(script, following)
        # Unfollow triggers a confirmation dialog — click the confirm button
        if isinstance(result, dict) and result.get("needsConfirm"):
            time.sleep(1)
            self.driver.execute_script(r'''
            var modal = document.querySelector('[role="dialog"], [role="alertdialog"], .artdeco-modal');
            if (modal) {
                var btns = modal.querySelectorAll('button');
                for (var i = 0; i < btns.length; i++) {
                    var cls = btns[i].className || '';
                    if (cls.indexOf('artdeco-button--primary') >= 0) {
                        btns[i].click();
                        return;
                    }
                }
                // Fallback: click last button (usually the confirm action)
                if (btns.length > 1) btns[btns.length - 1].click();
            }
            ''')
            time.sleep(1)
        return result

    def unfollow_entity(self, urn_id: str):
        """Unfollow an entity."""
        return self.follow_company(public_id=urn_id, following=False)

    def get_job(self, job_id: str) -> dict:
        data = self._api_get(f"/jobs/jobPostings/{job_id}")
        # Scrape company name from job page if not resolved
        company_details = data.get("companyDetails", {})
        if isinstance(company_details, dict):
            inner = company_details.get("com.linkedin.voyager.jobs.JobPostingCompany", company_details)
            if not inner.get("companyResolutionResult"):
                try:
                    self.driver.get(f"https://www.linkedin.com/jobs/view/{job_id}/")
                    time.sleep(3)
                    name = self.driver.execute_script(
                        r'''var links = document.querySelectorAll('a[href*="/company/"]');
                        for (var i = 0; i < links.length; i++) {
                            var t = links[i].innerText.trim().split('\n')[0].trim();
                            if (t) return t;
                        }
                        return '';'''
                    )
                    if name:
                        inner["companyResolutionResult"] = {"name": name}
                except Exception:
                    pass
        return data

    def get_job_skills(self, job_id: str) -> list:
        """Get skills required for a job."""
        # Try API first
        data = self._api_get(f"/jobs/jobPostings/{job_id}/skillMatchStatuses")
        if data and data.get("elements"):
            return data.get("elements", [])
        # Fallback: extract skills from the job description via API
        job = self._api_get(f"/jobs/jobPostings/{job_id}")
        desc_text = ""
        desc = job.get("description", {})
        if isinstance(desc, dict):
            desc_text = desc.get("text", "")
        elif isinstance(desc, str):
            desc_text = desc
        if not desc_text:
            return []
        # Extract qualifications/skills from description text
        import re
        skills = []
        seen = set()
        skip_words = ['equal opportunity', 'accommodat', 'position will be open',
                       'microsoft is', 'benefits', 'salary', 'compensation']
        # Split into sections by common headers (multi-language)
        header_pattern = (
            r'Required|Preferred|Qualifications|Skills|Requirements'  # EN
            r'|Kenntnisse|Anforderungen|Voraussetzungen|Qualifikationen'  # DE
            r'|Comp[eé]tences|Exigences|Pr[eé]requis|Qualifications'  # FR
            r'|Competenze|Requisiti|Qualifiche'  # IT
            r'|Habilidades|Requisitos|Cualificaciones|Competencias'  # ES
            r'|Vaardigheden|Vereisten|Kwalificaties'  # NL
        )
        sections = re.split(
            r'\n(?=(?:' + header_pattern + r'))',
            desc_text, flags=re.IGNORECASE
        )
        for section in sections:
            header_match = re.match(
                r'(' + header_pattern + r')[^\n]*',
                section, re.IGNORECASE
            )
            if not header_match:
                continue
            body = section[header_match.end():].strip()
            if not body:
                continue
            # Split by sentence boundaries or line breaks
            items = re.split(r'(?:\.\s*(?=[A-Z0-9])|\n)', body)
            for item in items:
                item = re.sub(r'^[\s•\-\*\d.)+]+', '', item).strip()
                if not item or len(item) < 10 or len(item) > 200:
                    continue
                if any(sw in item.lower() for sw in skip_words):
                    continue
                if item.lower() not in seen:
                    seen.add(item.lower())
                    skills.append({"name": item})
        return skills

    def search_companies(self, keywords=None, limit=25) -> list:
        kw = keywords[0] if isinstance(keywords, list) else (keywords or "")
        if not kw:
            return []
        results = []
        page_size = min(limit, 49)
        start = 0
        while start < limit:
            count = min(page_size, limit - start)
            data = self._api_get(
                f"/search/dash/clusters?decorationId=com.linkedin.voyager.dash.deco.search.SearchClusterCollection-175"
                f"&origin=SWITCH_SEARCH_VERTICAL&q=all&start={start}&count={count}"
                f"&query=(keywords:{kw},flagshipSearchIntent:SEARCH_SRP"
                f",queryParameters:(resultType:List(COMPANIES)))"
            )
            page = self._extract_search_results(data)
            results.extend(page)
            if len(page) < count:
                break
            start += len(page)
        return results[:limit]

    def search_posts(self, keywords=None, limit=25, sort_by=None,
                     date_posted=None, content_type=None,
                     from_member=None, from_company=None,
                     posted_by=None, mentioning=None) -> list:
        """Search for posts/content on LinkedIn with filters.

        Filters:
            sort_by: "relevance" or "date_posted"
            date_posted: "past-24h", "past-week", "past-month"
            content_type: "videos", "images", "documents", "job-postings", "liveVideos"
            from_member: member URN (ACoAAB...)
            from_company: company URN or ID
            posted_by: "first" (1st connections), "following" (people you follow)
            mentioning: member URN
        """
        import re as _re
        from urllib.parse import quote
        kw = quote(keywords or "")
        if not kw:
            return []

        # Build filter URL params
        params = [f"keywords={kw}", "origin=FACETED_SEARCH"]
        if sort_by and sort_by != "relevance":
            params.append(f"sortBy=%5B%22{quote(sort_by)}%22%5D")
        if date_posted:
            params.append(f"datePosted=%5B%22{quote(date_posted)}%22%5D")
        if content_type:
            params.append(f"contentType=%5B%22{quote(content_type)}%22%5D")
        if from_member:
            params.append(f"fromMember=%5B%22{quote(from_member)}%22%5D")
        if from_company:
            params.append(f"fromOrganization=%5B%22{quote(from_company)}%22%5D")
        if posted_by:
            params.append(f"postedBy=%5B%22{quote(posted_by)}%22%5D")
        if mentioning:
            params.append(f"mentionedMember=%5B%22{quote(mentioning)}%22%5D")

        all_posts = []
        seen = set()
        pages_needed = max(1, (limit + 2) // 3)  # ~3 posts per page

        for page in range(1, pages_needed + 1):
            url = f"https://www.linkedin.com/search/results/content/?{'&'.join(params)}&page={page}"
            self.driver.get(url)
            time.sleep(6)

            # Extract activity URNs from page source (order matches DOM)
            source = self.driver.page_source
            urn_list = []
            for m in _re.finditer(r'urn:li:activity:(\d+)', source):
                if m.group(1) not in seen and m.group(1) not in [u for u in urn_list]:
                    urn_list.append(m.group(1))

            # Extract posts from DOM
            page_posts = self.driver.execute_script(r'''
            var containers = document.querySelectorAll('[data-view-name="feed-full-update"]');
            var posts = [];
            for (var ci = 0; ci < containers.length; ci++) {
                var el = containers[ci];
                var text = el.innerText || "";
                var lines = text.split("\n").map(function(l) { return l.trim(); }).filter(function(l) { return l; });
                var startIdx = 0;
                for (var j = 0; j < lines.length; j++) {
                    if (lines[j].match(/^(Feed.post|Feed.Beitrag|Nummer)/i)) { startIdx = j + 1; continue; }
                    break;
                }
                while (startIdx < lines.length && lines[startIdx].match(/^(Vorgeschlagen|Suggested)/i)) startIdx++;
                var author = startIdx < lines.length ? lines[startIdx] : "";
                author = author.replace(/\s*(Verified|Verifiziert|Premium|Open to work|Offen f).*/i, "")
                    .replace(/\s+\d+(st|nd|rd|th|\.)?\+?\s*$/i, "").trim();
                var postText = "";
                for (var j = startIdx + 1; j < lines.length; j++) {
                    if (lines[j].length > postText.length && lines[j].length > 20 &&
                        !lines[j].match(/^\d+\s*(Reaktion|reaction|Kommentar|comment|Repost|Share)/i) &&
                        !lines[j].match(/^(Follow|Folgen|Melden|Report|Alle\s\d|Zur\sWebsite|To\swebsite)/i)) {
                        postText = lines[j];
                    }
                }
                var rxMatch = text.match(/(\d[\d.,]*)\s*(Reaktion|reaction)/i);
                var cmMatch = text.match(/(\d[\d.,]*)\s*(Kommentar|comment)/i);
                posts.push({
                    author: author, text: postText.substring(0, 500),
                    reactions: rxMatch ? rxMatch[1] : "0",
                    comments: cmMatch ? cmMatch[1] : "0"
                });
            }
            return posts;
            ''') or []

            if not page_posts:
                break

            for i, p in enumerate(page_posts):
                activity_id = urn_list[i] if i < len(urn_list) else ""
                if activity_id in seen:
                    continue
                seen.add(activity_id)
                urn = f"urn:li:activity:{activity_id}" if activity_id else ""
                all_posts.append({
                    "author": p["author"],
                    "text": p["text"],
                    "reactions": p["reactions"],
                    "comments": p["comments"],
                    "activity_id": activity_id,
                    "urn": urn,
                    "posted_at": _urn_to_timestamp(activity_id) if activity_id else "",
                    "post_url": f"https://www.linkedin.com/feed/update/{urn}/" if urn else "",
                })

            if len(all_posts) >= limit:
                break

        return all_posts[:limit]

    def search_groups(self, keywords=None, limit=25) -> list:
        """Search for groups on LinkedIn."""
        kw = keywords or ""
        if not kw:
            return []
        results = []
        page_size = min(limit, 49)
        start = 0
        while start < limit:
            count = min(page_size, limit - start)
            data = self._api_get(
                f"/search/dash/clusters?decorationId=com.linkedin.voyager.dash.deco.search.SearchClusterCollection-175"
                f"&origin=SWITCH_SEARCH_VERTICAL&q=all&start={start}&count={count}"
                f"&query=(keywords:{kw},flagshipSearchIntent:SEARCH_SRP"
                f",queryParameters:(resultType:List(GROUPS)))"
            )
            page = self._extract_search_results(data)
            results.extend(page)
            if len(page) < count:
                break
            start += len(page)
        return results[:limit]

    def search_events(self, keywords=None, limit=25) -> list:
        """Search for events on LinkedIn."""
        kw = keywords or ""
        if not kw:
            return []
        results = []
        page_size = min(limit, 49)
        start = 0
        while start < limit:
            count = min(page_size, limit - start)
            data = self._api_get(
                f"/search/dash/clusters?decorationId=com.linkedin.voyager.dash.deco.search.SearchClusterCollection-175"
                f"&origin=SWITCH_SEARCH_VERTICAL&q=all&start={start}&count={count}"
                f"&query=(keywords:{kw},flagshipSearchIntent:SEARCH_SRP"
                f",queryParameters:(resultType:List(EVENTS)))"
            )
            page = self._extract_search_results(data)
            results.extend(page)
            if len(page) < count:
                break
            start += len(page)
        return results[:limit]

    def search_jobs(self, keywords=None, limit=25, **kwargs) -> list:
        """Search jobs by scraping the LinkedIn jobs search page."""
        from urllib.parse import quote
        kw = quote(keywords or "")
        location = quote(kwargs.get("location_name", ""))
        base_url = f"https://www.linkedin.com/jobs/search/?keywords={kw}"
        if location:
            base_url += f"&location={location}"

        all_results = []
        seen = {}
        page_size = 25  # LinkedIn shows 25 jobs per page
        start = 0

        scrape_script = r'''
        var results = [];
        var seen = arguments[0];
        var items = document.querySelectorAll('.scaffold-layout__list-item, [data-job-id]');
        for (var i = 0; i < items.length; i++) {
            var card = items[i];
            // Find data-job-id on this element or inside it
            var jobId = card.getAttribute('data-job-id') || '';
            if (!jobId) {
                var inner = card.querySelector('[data-job-id]');
                if (inner) jobId = inner.getAttribute('data-job-id');
            }
            if (!jobId || seen[jobId]) continue;
            seen[jobId] = true;
            var titleEl = card.querySelector('.job-card-list__title, .artdeco-entity-lockup__title, a[class*="job-card"]');
            var companyEl = card.querySelector('.job-card-container__primary-description, .artdeco-entity-lockup__subtitle');
            var locationEl = card.querySelector('.job-card-container__metadata-item, .artdeco-entity-lockup__caption');
            if (!titleEl) continue;
            results.push({
                name: titleEl ? titleEl.innerText.trim().split('\n')[0] : '',
                headline: companyEl ? companyEl.innerText.trim() : '',
                location: locationEl ? locationEl.innerText.trim() : '',
                urn_id: 'urn:li:jobPosting:' + jobId
            });
        }
        return {results: results, seen: seen};
        '''

        while start < limit:
            url = f"{base_url}&start={start}"
            self.driver.get(url)
            time.sleep(4)

            # Scroll each list item into view to trigger LinkedIn's lazy rendering
            item_count = self.driver.execute_script(
                "return document.querySelectorAll('.scaffold-layout__list-item').length;"
            ) or 0
            for i in range(item_count):
                self.driver.execute_script(
                    "var items = document.querySelectorAll('.scaffold-layout__list-item');"
                    f"if (items[{i}]) items[{i}].scrollIntoView({{block: 'center'}});"
                )
                time.sleep(0.3)
            time.sleep(1)

            page_data = self.driver.execute_script(scrape_script, seen) or {}
            page_results = page_data.get("results", [])
            seen = page_data.get("seen", seen)
            all_results.extend(page_results)

            if len(page_results) < page_size // 2:
                break  # no more pages
            start += page_size

        return all_results[:limit]

    def get_post_analytics(self, post_urn: str) -> dict:
        """Get post analytics via GraphQL export endpoint. Returns performance metrics and demographics."""
        import re as _re, base64, tempfile, os

        activity_id = post_urn.split(":")[-1]
        # Load the analytics page to capture the queryId
        self.driver.get(f"https://www.linkedin.com/analytics/post-summary/urn:li:activity:{activity_id}/")
        time.sleep(4)

        qid_url = self.driver.execute_script(r'''
        var entries = performance.getEntriesByType("resource");
        for (var i = 0; i < entries.length; i++) {
            if (entries[i].name.indexOf("AnalyticsExports") >= 0) return entries[i].name;
        }
        return null;
        ''')

        if not qid_url:
            # Try hardcoded queryId
            qid = "voyagerPremiumDashAnalyticsExports.010b8b1d4bf6998ddff7ffecc03aa938"
        else:
            qid_match = _re.search(r'queryId=([^&]+)', qid_url)
            qid = qid_match.group(1) if qid_match else ""

        if not qid:
            return {}

        view_urn = f"urn%3Ali%3Afsd_edgeInsightsAnalyticsView%3A%28POST_SUMMARY%2Curn%3Ali%3Aactivity%3A{activity_id}%29"
        data = self._api_get(f"/graphql?variables=(analyticsViewUrn:{view_urn})&queryId={qid}")
        feed = (data or {}).get("data", {})

        # Find download URL
        dl_url = None
        for k, v in feed.items():
            if isinstance(v, dict) and "elements" in v:
                for el in v["elements"]:
                    if "downloadUrl" in el:
                        dl_url = el["downloadUrl"]
                        break

        if not dl_url:
            return {}

        # Download XLSX via browser
        b64 = self.driver.execute_script(r'''
        var resp = await fetch(arguments[0], {credentials: "include"});
        var buf = await resp.arrayBuffer();
        var bytes = new Uint8Array(buf);
        var binary = '';
        for (var i = 0; i < bytes.length; i++) binary += String.fromCharCode(bytes[i]);
        return btoa(binary);
        ''', dl_url)

        if not b64:
            return {}

        # Parse XLSX
        xlsx_path = os.path.join(tempfile.gettempdir(), f"post_analytics_{activity_id}.xlsx")
        with open(xlsx_path, "wb") as f:
            f.write(base64.b64decode(b64))

        try:
            from openpyxl import load_workbook
        except ImportError:
            return {"error": "openpyxl not installed — run: pip install openpyxl"}

        wb = load_workbook(xlsx_path)
        result = {"activity_id": activity_id}

        # Parse PERFORMANCE sheet
        if "PERFORMANCE" in wb.sheetnames:
            ws = wb["PERFORMANCE"]
            for row in ws.iter_rows(values_only=True):
                if row[0] and row[1]:
                    key = str(row[0]).strip()
                    val = str(row[1]).strip()
                    if key in ("Impressions", "Reactions", "Comments", "Reposts", "Saves",
                               "Members reached", "Profile viewers from this post",
                               "Followers gained from this post", "Sends on LinkedIn"):
                        result[key] = val
                    elif key == "Post Date":
                        result["date"] = val
                    elif key == "Post URL":
                        result["url"] = val

        # Parse TOP DEMOGRAPHICS sheet
        if "TOP DEMOGRAPHICS" in wb.sheetnames:
            ws = wb["TOP DEMOGRAPHICS"]
            demographics = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    category = str(row[0]).strip()
                    value = str(row[1]).strip()
                    pct = row[2]
                    if category not in demographics:
                        demographics[category] = []
                    pct_str = f"{pct:.1%}" if isinstance(pct, (int, float)) else str(pct or "")
                    demographics[category].append({"value": value, "pct": pct_str})
            result["demographics"] = demographics

        os.remove(xlsx_path)
        return result

    def get_post(self, post_urn: str) -> dict:
        """Get detailed post data via REST API."""
        activity_id = post_urn.split(":")[-1]
        data = self._api_get(f"/feed/updates/urn:li:activity:{activity_id}")
        if not data:
            return {}

        update = data.get("value", {}).get("com.linkedin.voyager.feed.render.UpdateV2", {})

        actor = update.get("actor") or {}
        name = (actor.get("name") or {}).get("text", "")
        desc = (actor.get("description") or {}).get("text", "")
        date = (actor.get("subDescription") or {}).get("text", "")

        commentary = update.get("commentary") or {}
        text = (commentary.get("text") or {}).get("text", "")

        social = update.get("socialDetail") or {}
        counts = social.get("totalSocialActivityCounts") or {}

        # Content type
        content = update.get("content") or {}
        content_type = ""
        for k in content.keys():
            if "Video" in k:
                content_type = "Video"
            elif "Image" in k:
                content_type = "Image"
            elif "Article" in k:
                content_type = "Article"
            elif "Document" in k:
                content_type = "Document"

        return {
            "urn": f"urn:li:activity:{activity_id}",
            "url": data.get("permalink", ""),
            "author": name,
            "headline": desc,
            "date": date,
            "text": text,
            "content_type": content_type,
            "impressions": counts.get("numImpressions", 0),
            "views": counts.get("numViews", 0),
            "reactions": counts.get("numLikes", 0),
            "comments": counts.get("numComments", 0),
            "shares": counts.get("numShares", 0),
        }

    def get_notifications(self, limit=25, unread_only=False) -> list:
        """Get notifications via Voyager REST API with pagination."""
        from datetime import datetime
        all_notifs = []
        page_size = 20
        start = 0

        while start < limit:
            count = min(page_size, limit - start)
            data = self._api_get(
                f"/voyagerIdentityDashNotificationCards?count={count}&start={start}&q=filterVanityName"
            )
            elements = (data or {}).get("elements", [])
            if not elements:
                break

            for el in elements:
                read = el.get("read", True)
                if unread_only and read:
                    continue

                headline = el.get("headline") or {}
                h_text = headline.get("text", "") if isinstance(headline, dict) else str(headline)

                ts = el.get("publishedAt", 0)
                date_str = datetime.fromtimestamp(ts / 1000).strftime("%Y-%m-%d %H:%M") if ts else ""

                # Extract notification URN and action URL
                notif_urn = el.get("entityUrn", "")
                action_url = el.get("navigationUrl", "")

                # Extract actor info if available
                actor_name = ""
                actor_profile_id = ""
                actors = el.get("actors", [])
                if actors and isinstance(actors, list):
                    first_actor = actors[0] if actors else {}
                    if isinstance(first_actor, dict):
                        a_name = first_actor.get("name") or {}
                        actor_name = a_name.get("text", "") if isinstance(a_name, dict) else str(a_name)
                        a_nav = first_actor.get("navigationUrl", "")
                        if "/in/" in a_nav:
                            actor_profile_id = a_nav.split("/in/")[-1].split("?")[0].rstrip("/")

                all_notifs.append({
                    "headline": h_text,
                    "date": date_str,
                    "read": read,
                    "notification_urn": notif_urn,
                    "action_url": action_url,
                    "actor_name": actor_name,
                    "actor_profile_id": actor_profile_id,
                })

            start += len(elements)
            if len(elements) < count:
                break

        return all_notifs[:limit]

    def get_post_engagers(self, post_urn: str, limit: int = 50) -> list:
        """Get combined reactions + comments with deduplication by profileId."""
        if not post_urn.startswith("urn:"):
            post_urn = f"urn:li:activity:{post_urn}"

        reactions = self.get_post_reactions(post_urn=post_urn, limit=limit)
        comments = self.get_post_comments(post_urn=post_urn, limit=limit)

        seen = {}
        merged = []

        for r in reactions:
            pid = r.get("profileId", "")
            if pid and pid not in seen:
                seen[pid] = len(merged)
                merged.append({
                    "name": r.get("name", ""),
                    "headline": r.get("headline", ""),
                    "profileId": pid,
                    "profileUrl": r.get("profileUrl", ""),
                    "interaction_type": "reaction",
                })

        for c in comments:
            pid = c.get("profileId", "")
            if pid and pid in seen:
                merged[seen[pid]]["interaction_type"] = "both"
            elif pid:
                seen[pid] = len(merged)
                merged.append({
                    "name": c.get("author", ""),
                    "headline": "",
                    "profileId": pid,
                    "profileUrl": c.get("profileUrl", ""),
                    "interaction_type": "comment",
                })

        return merged[:limit]

    def get_signals(self, recent_post_urns: list = None, limit: int = 5) -> dict:
        """Aggregate daily signals: profile views, post engagers, invitations, notifications."""
        signals = {}

        # Profile views
        signals["profile_views"] = self.get_current_profile_views()

        # Post engagers for recent posts
        engagers = []
        for urn in (recent_post_urns or [])[:3]:
            post_engagers = self.get_post_engagers(post_urn=urn, limit=limit)
            for e in post_engagers:
                e["post_urn"] = urn
            engagers.extend(post_engagers)
        signals["post_engagers"] = engagers

        # Invitations
        signals["invitations"] = self.get_invitations(limit=limit)

        # Unread notifications
        signals["notifications"] = self.get_notifications(limit=limit, unread_only=True)

        return signals

    def quit(self):
        """Close the browser and clean up temp profile."""
        try:
            self.driver.quit()
        except Exception:
            pass
        temp = getattr(self, "_temp_profile", None)
        if temp and Path(str(temp)).exists():
            import shutil
            shutil.rmtree(temp, ignore_errors=True)


def _normalize_profile(raw: dict) -> dict:
    """Normalize dash profile response to standard format."""
    first = ""
    last = ""

    multi_first = raw.get("multiLocaleFirstName", {})
    if multi_first:
        first = next(iter(multi_first.values()), "")

    multi_last = raw.get("multiLocaleLastName", {})
    if multi_last:
        last = next(iter(multi_last.values()), "")

    headline = ""
    multi_headline = raw.get("multiLocaleHeadline", {})
    if multi_headline:
        headline = next(iter(multi_headline.values()), "")

    # Location from geoLocation or location fields
    location = raw.get("geoLocationName", "")
    if not location:
        loc = raw.get("location", {})
        geo = raw.get("geoLocation", {})
        parts = [loc.get("city", ""), geo.get("postalCode", ""), loc.get("countryCode", "").upper()]
        location = ", ".join(p for p in parts if p)

    # Follower count from followingState (with TopCard decoration)
    following_state = raw.get("followingState") or {}
    follower_count = following_state.get("followerCount", "")

    # Connection count from connections paging
    connections = raw.get("connections") or {}
    paging = connections.get("paging", {})
    connection_count = paging.get("total", "")

    # Summary from multiLocale or plain
    summary = ""
    multi_summary = raw.get("multiLocaleSummary", {})
    if multi_summary:
        summary = next(iter(multi_summary.values()), "")
    if not summary:
        summary = raw.get("summary", "")

    return {
        "firstName": first or raw.get("firstName", ""),
        "lastName": last or raw.get("lastName", ""),
        "headline": headline,
        "locationName": location,
        "industryName": raw.get("industryName", ""),
        "summary": summary,
        "followerCount": follower_count,
        "connectionCount": connection_count,
        "entityUrn": raw.get("entityUrn", ""),
        "objectUrn": raw.get("objectUrn", ""),
        "publicIdentifier": raw.get("publicIdentifier", ""),
        "_raw": raw,
    }
